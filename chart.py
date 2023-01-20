from openpyxl import chart
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import column_index_from_string, get_column_letter

# 라인 차트 그리기
def LineChart(a, b, c, d, e, f):
    data_col = a         # 데이터 범주
    cat_col = b          # 항목 범주
    title = c            # 차트 제목
    sheet = d            # 차트가 그려질 시트
    cell_val = e         # 차트가 그려질 cell값 (셀 숫자)
    chart_num = f        # 차트가 그려질 cell값 (차트 넘버)

    position = chart_posititon(cell_val, chart_num)

    LC = chart.LineChart()
    LC.title = title

    for i in data_col:
        col = column_index_from_string(i)
        data = chart.Reference(sheet, min_col=col, min_row=1,
                               max_col=col, max_row=sheet.max_row)
        LC.add_data(data, titles_from_data=True)

        for i in cat_col:
            cat = column_index_from_string(i)
            category = chart.Reference(sheet, min_col=cat, min_row=2,
                                      max_col=cat, max_row=sheet.max_row)
            LC.set_categories(category)

    LC.style = 2

    sheet.add_chart(LC, position)

def PieChart(a, b, c, d, e, f, g):
    data_col = a         # 데이터 범주
    cat_col = b          # 항목 범주
    title = c            # 차트 제목
    sheet = d            # 차트가 그려질 시트
    cell_val = e         # 차트가 그려질 cell값 (셀 숫자)
    chart_num = f        # 차트가 그려질 cell값 (차트 넘버)
    top = g              # TOP 순위 출력

    position = chart_posititon(cell_val, chart_num)

    LC = chart.PieChart()
    LC.title = title

    for i in data_col:
        col = column_index_from_string(i)
        data = chart.Reference(sheet, min_col=col, min_row=1,
                               max_col=col, max_row=top+1)
        LC.add_data(data, titles_from_data=True)
        LC.dataLabels = DataLabelList()


        for i in cat_col:
            cat = column_index_from_string(i)
            category = chart.Reference(sheet, min_col=cat, min_row=2,
                                      max_col=cat, max_row=top+1)
            LC.set_categories(category)

    LC.style = 2

    LC.dataLabels.showPercent = True
    LC.dataLabels.showLeaderLines = True

    sheet.add_chart(LC, position)

# 차트 좌표 구하기
def chart_posititon(a, b):
    max_cell_val = int(a)
    chart_num = int(b)

    cell_num = (chart_num // 2) * 15 + 2

    if chart_num % 2 == 0:
        cell_val = get_column_letter(max_cell_val)
        position = str(cell_val+str(cell_num))
    else:
        cell_val = get_column_letter(max_cell_val+9)
        position = str(cell_val + str(cell_num))
    return position