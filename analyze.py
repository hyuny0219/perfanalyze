from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import cx_Oracle
import os
import analyzeSQL
import analyzeSQL_RAC
import pandas as pd
from conndb import connectDB
from chart import LineChart, PieChart

def execAnalyzeRAC(a, b, c, d, e, f, g, h, i):
    conn_string = a
    conn_id = b
    conn_pwd = c
    vsnap_start = str(d)
    vsnap_end = str(e)
    vdbid = str(f)
    path = str(i)

    dbname = str(conn_string.split('/')[1])
    today = datetime.now()
    today = today.strftime("%Y%m%d%H%M%S")

    conn = connectDB(conn_string, conn_id, conn_pwd)
    cursor = conn.cursor()

    wb = Workbook()
    sheet = wb.active
    file_name = "Perf_RAC_" + dbname + "_" + today

    # [Step 00 LOAD_2NODE]
    sheet.title = 'load_2node'

    sql = analyzeSQL_RAC.load_2node
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    # TOP LOAD_2NODE 컬럼명 수정
    sql = analyzeSQL_RAC.load_2node2
    cursor.execute(sql, snap_end=vsnap_end, dbid=vdbid)

    replace_column = cursor.fetchall()

    if not replace_column:
        pass
    else:
        j = 0
        for i in range(3, len(columns)):
            columns[i] = replace_column[j][0]
            j = j + 1

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)
    ## 그래프가 있을 경우 작성
    max_cell_val = int(sheet.max_column) + 2

    # 0.Transaction(/sec)
    data_col = ["D", "E"]
    cat_col = ["C"]
    title = "Transactions(/sec)"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 1.User Rollbacks(/sec)
    data_col = ["BJ", "BK"]
    cat_col = ["C"]
    title = "User Rollbacks(/sec)"
    chart_num = 1
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 2.Redo Size(bytes/sec)
    data_col = ["F", "G"]
    cat_col = ["C"]
    title = "Redo Size(bytes/sec)"
    chart_num = 2
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 3.Redo NoWait(%)
    data_col = ["H", "I"]
    cat_col = ["C"]
    title = "Redo NoWait(%)"
    chart_num = 3
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 4.Logons(/sec)
    data_col = ["J", "K"]
    cat_col = ["C"]
    title = "Logons(/sec)"
    chart_num = 4
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 5.User calls(/sec)
    data_col = ["L", "M"]
    cat_col = ["C"]
    title = "User calls(/sec)"
    chart_num = 5
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 6.Executes(/sec)
    data_col = ["N", "O"]
    cat_col = ["C"]
    title = "Executes(/sec)"
    chart_num = 6
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 7.Block Changes(/sec)
    data_col = ["P", "Q"]
    cat_col = ["C"]
    title = "Block Changes(/sec)"
    chart_num = 7
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 8.Buffer Hit(%)
    data_col = ["R", "S"]
    cat_col = ["C"]
    title = "Buffer Hit(%)"
    chart_num = 8
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 9.Buffer Nowait(%)
    data_col = ["T", "U"]
    cat_col = ["C"]
    title = "Buffer Nowait(%)"
    chart_num = 9
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 10.Library Hit(%)
    data_col = ["V", "W"]
    cat_col = ["C"]
    title = "Library Hit(%)"
    chart_num = 10
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 11.Latch Hit(%)
    data_col = ["X", "Y"]
    cat_col = ["C"]
    title = "Latch Hit(%)"
    chart_num = 11
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 12.Sorts(/sec)
    data_col = ["Z", "AA", "AB", "AC"]
    cat_col = ["C"]
    title = "Sorts(/sec)"
    chart_num = 12
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 13.In Memory Sort(%)
    data_col = ["AD", "AE"]
    cat_col = ["C"]
    title = "In Memory Sort(%)"
    chart_num = 13
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 14.Logical/Physical Reads(/sec)
    data_col = ["AF", "AG", "AH", "AI"]
    cat_col = ["C"]
    title = "Logical/Physical Reads(/sec)"
    chart_num = 14
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 15.Physical Read IO(/sec)
    data_col = ["AH", "AI", "AJ", "AK"]
    cat_col = ["C"]
    title = "Physical Read IO(/sec)"
    chart_num = 15
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 16.Physical Write IO(/sec)
    data_col = ["AL", "AM", "AN", "AO"]
    cat_col = ["C"]
    title = "Physical Write IO(/sec)"
    chart_num = 16
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 17.Hard Parse(/sec)
    data_col = ["AR", "AS"]
    cat_col = ["C"]
    title = "Hard Parse(/sec)"
    chart_num = 17
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 18.Hard Parse(%)
    data_col = ["AT", "AU"]
    cat_col = ["C"]
    title = "Hard Parse(%)"
    chart_num = 18
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 19.Soft Parse(%)
    data_col = ["AV", "AW"]
    cat_col = ["C"]
    title = "Soft Parse(%)"
    chart_num = 19
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 20.Execute to Parse(%)
    data_col = ["AX", "AY"]
    cat_col = ["C"]
    title = "Execute to Parse(%)"
    chart_num = 20
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 21.Parse CPU to Parse Elapsd(%)
    data_col = ["AZ", "BA"]
    cat_col = ["C"]
    title = "Parse CPU to Parse Elapsd(%)"
    chart_num = 21
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 22.Non-Parse CPU(%)
    data_col = ["BB", "BC"]
    cat_col = ["C"]
    title = "Non-Parse CPU(%)"
    chart_num = 22
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 23.Parse CPU/Elapsed Time(/sec)
    data_col = ["BD", "BE", "BF", "BG"]
    cat_col = ["C"]
    title = "Parse CPU/Elapsed Time(/sec)"
    chart_num = 23
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 01 RAC_LOAD_2NODE]
    sheet = wb.create_sheet('rac_load_2node')
    wb.active = wb.sheetnames.index('rac_load_2node')
    sheet = wb.active

    sql = analyzeSQL_RAC.rac_load_2node
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    # TOP RAC_LOAD_2NODE 컬럼명 수정
    sql = analyzeSQL_RAC.rac_load_2node2
    cursor.execute(sql, snap_end=vsnap_end, dbid=vdbid)

    replace_column = cursor.fetchall()

    if not replace_column:
        pass
    else:
        j = 0
        for i in range(3, len(columns)):
            columns[i] = replace_column[j][0]
            j = j + 1

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)
    ## 그래프가 있을 경우 작성
    max_cell_val = int(sheet.max_column) + 2

    # 0.Interconnect Traffic(MB)
    data_col = ["D", "E"]
    cat_col = ["C"]
    title = "Interconnect Traffic(MB)"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 1.GC blocks Received(/sec)
    data_col = ["F", "G", "H", "I", "J", "K"]
    cat_col = ["C"]
    title = "GC blocks Received(/sec)"
    chart_num = 1
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 2.GC Block Served(/sec)
    data_col = ["L", "M", "N", "O", "P", "Q"]
    cat_col = ["C"]
    title = "GC Block Served(/sec)"
    chart_num = 2
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 3.GCS/GES messages(/sec)
    data_col = ["R", "S", "T", "U"]
    cat_col = ["C"]
    title = "GCS/GES messages(/sec)"
    chart_num = 3
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 4.DBWR fusion writes(/sec)
    data_col = ["V", "W"]
    cat_col = ["C"]
    title = "DBWR fusion writes(/sec)"
    chart_num = 4
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 5.GC Blocks lost(/sec)
    data_col = ["X", "Y"]
    cat_col = ["C"]
    title = "GC Blocks lost(/sec)"
    chart_num = 5
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 6.GC Blocks corrupt(/sec)
    data_col = ["Z", "AA"]
    cat_col = ["C"]
    title = "GC Blocks corrupt(/sec)"
    chart_num = 6
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 7.Node1-Buffer Access Efficiency(%)
    data_col = ["AB", "AD", "AF"]
    cat_col = ["C"]
    title = "Node1-Buffer Access Efficiency(%)"
    chart_num = 7
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 8.Node2-Buffer Access Efficiency(%)
    data_col = ["AC", "AE", "AG"]
    cat_col = ["C"]
    title = "Node2-Buffer Access Efficiency(%)"
    chart_num = 8
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 02 RAC_LOAD2_2NODE]
    sheet = wb.create_sheet('rac_load2_2node')
    wb.active = wb.sheetnames.index('rac_load2_2node')
    sheet = wb.active

    sql = analyzeSQL_RAC.rac_load2_2node
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    # TOP RAC_LOAD2_2NODE 컬럼명 수정
    sql = analyzeSQL_RAC.rac_load_2node2
    cursor.execute(sql, snap_end=vsnap_end, dbid=vdbid)

    replace_column = cursor.fetchall()

    if not replace_column:
        pass
    else:
        j = 0
        for i in range(3, len(columns)):
            columns[i] = replace_column[j][0]
            j = j + 1

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)
    ## 그래프가 있을 경우 작성
    max_cell_val = int(sheet.max_column) + 2

    # 0.Avg GC cr block received Time(ms)
    data_col = ["D", "E"]
    cat_col = ["C"]
    title = "Avg GC cr block received Time(ms)"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 1.Avg GC current block Received Time(ms)
    data_col = ["F", "G"]
    cat_col = ["C"]
    title = "Avg GC current block Received Time(ms)"
    chart_num = 1
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 2.Node1-Avg GC cr block Served Time(ms)
    data_col = ["H", "J", "L"]
    cat_col = ["C"]
    title = "Node1-Avg GC cr block Served Time(ms)"
    chart_num = 2
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 3.Node2-Avg GC cr block Served Time(ms)
    data_col = ["I", "K", "M"]
    cat_col = ["C"]
    title = "Node2-Avg GC cr block Served Time(ms)"
    chart_num = 3
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 4.Avg CR block Flush Time(ms)
    data_col = ["J", "K"]
    cat_col = ["C"]
    title = "Avg CR block Flush Time(ms)"
    chart_num = 4
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 5.Node1-Avg GC current block Served Time(ms)
    data_col = ["N", "P", "R"]
    cat_col = ["C"]
    title = "Node1-Avg GC current block Served Time(ms)"
    chart_num = 5
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 6.Node2-Avg GC current block Served Time(ms)
    data_col = ["O", "Q", "S"]
    cat_col = ["C"]
    title = "Node2-Avg GC current block Served Time(ms)"
    chart_num = 6
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 7.Avg current block Flush Time(ms)
    data_col = ["R", "S"]
    cat_col = ["C"]
    title = "Avg current block Flush Time(ms)"
    chart_num = 7
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 8.Avg GE get time(ms)
    data_col = ["T", "U"]
    cat_col = ["C"]
    title = "Avg GE get time(ms)"
    chart_num = 8
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 03 RAC_MESSAGE_2NODE]
    sheet = wb.create_sheet('rac_message_2node')
    wb.active = wb.sheetnames.index('rac_message_2node')
    sheet = wb.active

    sql = analyzeSQL_RAC.rac_message_2node
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    # TOP RAC_MESSAGE_2NODE 컬럼명 수정
    sql = analyzeSQL_RAC.rac_message_2node2
    cursor.execute(sql, snap_end=vsnap_end, dbid=vdbid)

    replace_column = cursor.fetchall()

    if not replace_column:
        pass
    else:
        j = 0
        for i in range(3, len(columns)):
            columns[i] = replace_column[j][0]
            j = j + 1

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)
    ## 그래프가 있을 경우 작성
    max_cell_val = int(sheet.max_column) + 2

    # 0.Avg Message sent queue time(ms)
    data_col = ["D", "E"]
    cat_col = ["C"]
    title = "Avg Message sent queue time(ms)"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 1.Avg Message sent Queue time on ksxp(ms)
    data_col = ["F", "G"]
    cat_col = ["C"]
    title = "Avg Message sent Queue time on ksxp(ms)"
    chart_num = 1
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 2.Avg Message received queue time(ms)
    data_col = ["H", "I"]
    cat_col = ["C"]
    title = "Avg Message received queue time(ms)"
    chart_num = 2
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 3.Avg GCS message process time(ms)
    data_col = ["J", "K"]
    cat_col = ["C"]
    title = "Avg GCS message process time(ms)"
    chart_num = 3
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 4.Avg GES message process time(ms)
    data_col = ["L", "M"]
    cat_col = ["C"]
    title = "Avg GES message process time(ms)"
    chart_num = 4
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 5.Node1-sent Message ratio(%)
    data_col = ["N", "P", "R"]
    cat_col = ["C"]
    title = "Node1-sent Message ratio(%)"
    chart_num = 5
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 6.Node2-sent Message ratio(%)
    data_col = ["O", "Q", "S"]
    cat_col = ["C"]
    title = "Node2-sent Message ratio(%)"
    chart_num = 6
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 7.Node1-sent Messages
    data_col = ["T", "V", "X"]
    cat_col = ["C"]
    title = "Node1-sent Messages"
    chart_num = 7
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 8.Node2-sent Messages
    data_col = ["U", "W", "Y"]
    cat_col = ["C"]
    title = "Node2-sent Messages"
    chart_num = 8
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 04 OSSTAT_2NODE]
    sheet = wb.create_sheet('osstat_2node')
    wb.active = wb.sheetnames.index('osstat_2node')
    sheet = wb.active

    sql = analyzeSQL_RAC.osstat_2node
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    # TOP OSSTAT_2NODE 컬럼명 수정
    sql = analyzeSQL_RAC.osstat_2node2
    cursor.execute(sql, snap_end=vsnap_end, dbid=vdbid)

    replace_column = cursor.fetchall()

    if not replace_column:
        pass
    else:
        j = 0
        for i in range(3, len(columns)):
            columns[i] = replace_column[j][0]
            j = j + 1

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)
    ## 그래프가 있을 경우 작성
    max_cell_val = int(sheet.max_column) + 2

    # 0.Node1-CPU Usage(%)
    data_col = ["L", "N", "P", "R"]
    cat_col = ["C"]
    title = "Node1-CPU Usage(%)"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 1.Node2-CPU Usage(%)
    data_col = ["M", "O", "Q", "S"]
    cat_col = ["C"]
    title = "Node2-CPU Usage(%)"
    chart_num = 1
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 2.CPU Usage(%)
    data_col = ["T", "U"]
    cat_col = ["C"]
    title = "CPU Usage(%)"
    chart_num = 2
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 05 TIME_MODEL_2NODE]
    sheet = wb.create_sheet('time_model_2node')
    wb.active = wb.sheetnames.index('time_model_2node')
    sheet = wb.active

    sql = analyzeSQL_RAC.time_model_2node
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    # TOP TIME_MODEL_2NODE 컬럼명 수정
    sql = analyzeSQL_RAC.time_model_2node2
    cursor.execute(sql, snap_end=vsnap_end, dbid=vdbid)

    replace_column = cursor.fetchall()

    if not replace_column:
        pass
    else:
        j = 0
        for i in range(3, len(columns)):
            columns[i] = replace_column[j][0]
            j = j + 1

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)
    ## 그래프가 있을 경우 작성
    max_cell_val = int(sheet.max_column) + 2

    # 0.DB Time(s)
    data_col = ["D", "E", "F", "G"]
    cat_col = ["C"]
    title = "DB Time(s)"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    try:
        wb.save(path + "\\" + file_name + ".xlsx")
        result = "RAC 지표 생성 완료"
        return result
    except cx_Oracle.DatabaseError as err:
        return str(err)
    finally:
        cursor.close()
        conn.close()

def execAnalyze(a, b, c, d, e, f, g, h, i):
    conn_string = a
    conn_id = b
    conn_pwd = c
    vsnap_start = str(d)
    vsnap_end = str(e)
    vdbid = str(f)
    vinstno = str(g)
    instnm = h
    path = str(i)

    today = datetime.now()
    today = today.strftime("%Y%m%d%H%M%S")
    conn = connectDB(conn_string, conn_id, conn_pwd)
    cursor = conn.cursor()

    wb = Workbook()
    sheet = wb.active
    file_name = "Perf_"+ instnm + "_" + today

    # [Step 00 GEN_RPT]
    sheet.title = 'gen_rpt'

    sql = analyzeSQL.gen_rpt
    cursor.execute(sql)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)
    ## 그래프가 있을 경우 작성

    # [Step 01 DBINFO]
    sheet = wb.create_sheet('dbinfo')
    wb.active = wb.sheetnames.index('dbinfo')
    sheet = wb.active

    sql = analyzeSQL.dbinfo
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)
    ## 그래프가 있을 경우 작성

    # [Step 02 AWRSNAP]
    sheet = wb.create_sheet('awrsnap')
    wb.active = wb.sheetnames.index('awrsnap')
    sheet = wb.active

    sql = analyzeSQL.awrsnap
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    # [Step 03 LOAD]
    sheet = wb.create_sheet('load')
    wb.active = wb.sheetnames.index('load')
    sheet = wb.active

    sql = analyzeSQL.load
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    ## 그래프가 있을 경우 작성
    max_cell_val = int(sheet.max_column) + 2

    # 0.Transaction(/sec)
    data_col = ["D"]
    cat_col = ["C"]
    title = "Transactions(/sec)"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 1.Redo Size(bytes/sec)
    data_col = ["E"]
    cat_col = ["C"]
    title = "Redo Size(bytes/sec)"
    chart_num = 1
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 2.Redo NoWait(%)
    data_col = ["F"]
    cat_col = ["C"]
    title = "Redo NoWait(%)"
    chart_num = 2
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 3.Logons(/sec)
    data_col = ["G"]
    cat_col = ["C"]
    title = "Logons(/sec)"
    chart_num = 3
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 4.User calls(/sec)
    data_col = ["H"]
    cat_col = ["C"]
    title = "User calls(/sec)"
    chart_num = 4
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 5.Executes(/sec)
    data_col = ["I"]
    cat_col = ["C"]
    title = "Executes(/sec)"
    chart_num = 5
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 6.Block Changes(/sec)
    data_col = ["J"]
    cat_col = ["C"]
    title = "Block Changes(/sec)"
    chart_num = 6
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 7.Buffer Hit(%)
    data_col = ["K"]
    cat_col = ["C"]
    title = "Buffer Hit(%)"
    chart_num = 7
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 8.Buffer Nowait(%)
    data_col = ["L"]
    cat_col = ["C"]
    title = "Buffer Nowait(%)"
    chart_num = 8
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 9.Library Hit(%)
    data_col = ["M"]
    cat_col = ["C"]
    title = "Library Hit(%)"
    chart_num = 9
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 10.Latch Hit(%)
    data_col = ["N"]
    cat_col = ["C"]
    title = "Latch Hit(%)"
    chart_num = 10
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 11.Sorts(/sec)
    data_col = ["O", "p"]
    cat_col = ["C"]
    title = "Sorts(/sec)"
    chart_num = 11
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 12.In Memory Sort(%)
    data_col = ["Q"]
    cat_col = ["C"]
    title = "In Memory Sort(%)"
    chart_num = 12
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 13.Logical/Physical Reads(/sec)
    data_col = ["R", "S"]
    cat_col = ["C"]
    title = "Logical/Physical Reads(/sec)"
    chart_num = 13
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 14.Physical Read IO(/sec)
    data_col = ["S", "T"]
    cat_col = ["C"]
    title = "Physical Read IO(/sec)"
    chart_num = 14
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 15.Physical Write IO(/sec)
    data_col = ["U", "V"]
    cat_col = ["C"]
    title = "Physical Write IO(/sec)"
    chart_num = 15
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 16.Hard Parse(/sec)
    data_col = ["X"]
    cat_col = ["C"]
    title = "Hard Parse(/sec)"
    chart_num = 16
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 17.Hard Parse(%)
    data_col = ["Y"]
    cat_col = ["C"]
    title = "Hard Parse(%)"
    chart_num = 17
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 18.Soft Parse(%)
    data_col = ["Z"]
    cat_col = ["C"]
    title = "Soft Parse(%)"
    chart_num = 18
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 19.Execute to Parse(%)
    data_col = ["AA"]
    cat_col = ["C"]
    title = "Execute to Parse(%)"
    chart_num = 19
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 20.Parse CPU to Parse Elapsd(%)
    data_col = ["AB"]
    cat_col = ["C"]
    title = "Parse CPU to Parse Elapsd(%)"
    chart_num = 20
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 21.Non-Parse CPU(%)
    data_col = ["AC"]
    cat_col = ["C"]
    title = "Non-Parse CPU(%)"
    chart_num = 21
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 22.Parse CPU/Elapsed Time(/sec)
    data_col = ["AD", "AE"]
    cat_col = ["C"]
    title = "Parse CPU/Elapsed Time(/sec)"
    chart_num = 22
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 04 RAC_LOAD1]
    sheet = wb.create_sheet('rac_load1')
    wb.active = wb.sheetnames.index('rac_load1')
    sheet = wb.active

    sql = analyzeSQL.rac_load
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    ## 그래프가 있을 경우 작성
    max_cell_val = int(sheet.max_column) + 2

    # 0.Interconnect Traffic(MB)
    data_col = ["D"]
    cat_col = ["C"]
    title = "Interconnect Traffic(MB)"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 1.Traffic:received(MB)
    data_col = ["S"]
    cat_col = ["C"]
    title = "Traffic:received(MB)"
    chart_num = 1
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 2.Traffic:sent(MB)
    data_col = ["T"]
    cat_col = ["C"]
    title = "Traffic:sent(MB)"
    chart_num = 2
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 3.GC Block Received(/sec)
    data_col = ["E", "F", "G"]
    cat_col = ["C"]
    title = "GC Block Received(/sec)"
    chart_num = 3
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 4.GC Block Served(/sec)
    data_col = ["H", "I", "J"]
    cat_col = ["C"]
    title = "GC Block Served(/sec)"
    chart_num = 4
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 5.GCS/GES messages(/sec)
    data_col = ["K", "L"]
    cat_col = ["C"]
    title = "GCS/GES messages(/sec)"
    chart_num = 5
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 6.DBWR fusion writes(/sec)
    data_col = ["M"]
    cat_col = ["C"]
    title = "DBWR fusion writes(/sec)"
    chart_num = 6
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 7.GC Blocks lost(/sec)
    data_col = ["N"]
    cat_col = ["C"]
    title = "GC Blocks lost(/sec)"
    chart_num = 7
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 8.GC Blocks corrupt(/sec)
    data_col = ["O"]
    cat_col = ["C"]
    title = "GC Blocks corrupt(/sec)"
    chart_num = 8
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 9.Buffer Access Efficiency(%)
    data_col = ["P", "Q", "R"]
    cat_col = ["C"]
    title = "Buffer Access Efficiency(%)"
    chart_num = 9
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 05 RAC_LOAD2]
    sheet = wb.create_sheet('rac_load2')
    wb.active = wb.sheetnames.index('rac_load2')
    sheet = wb.active

    sql = analyzeSQL.rac_load2
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    ## 그래프가 있을 경우 작성
    max_cell_val = int(sheet.max_column) + 2

    # 0.Avg GC cr block received Time(ms)
    data_col = ["D"]
    cat_col = ["C"]
    title = "Avg GC cr block received Time(ms)"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 1.Avg GC current block Received Time(ms)
    data_col = ["E"]
    cat_col = ["C"]
    title = "Avg GC current block Received Time(ms)"
    chart_num = 1
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 2.Avg GC cr block Served Time(ms)
    data_col = ["F", "G", "H"]
    cat_col = ["C"]
    title = "Avg GC cr block Served Time(ms)"
    chart_num = 2
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 3. Avg GC current block Served time(ms)
    data_col = ["J", "K", "L"]
    cat_col = ["C"]
    title = "Avg GC current block Served time(ms)"
    chart_num = 3
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 4.Avg GE get time(ms)
    data_col = ["N"]
    cat_col = ["C"]
    title = "Avg GE get time(ms)"
    chart_num = 4
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 06 RAC_MESSAGE]
    sheet = wb.create_sheet('rac_message')
    wb.active = wb.sheetnames.index('rac_message')
    sheet = wb.active

    sql = analyzeSQL.rac_message
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    ## 그래프가 있을 경우 작성
    max_cell_val = int(sheet.max_column) + 2

    # 0.Avg Message sent queue time(ms)
    data_col = ["D"]
    cat_col = ["C"]
    title = "Avg Message sent queue time(ms)"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 1.Avg Message sent Queue time on ksxp(ms)
    data_col = ["E"]
    cat_col = ["C"]
    title = "Avg Message sent Queue time on ksxp(ms)"
    chart_num = 1
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 2.Avg Message received queue time(ms)
    data_col = ["F"]
    cat_col = ["C"]
    title = "Avg Message received queue time(ms)"
    chart_num = 2
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 3.Avg GCS message process time(ms)
    data_col = ["G"]
    cat_col = ["C"]
    title = "Avg GCS message process time(ms)"
    chart_num = 3
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 4.Avg GES message process time(ms)
    data_col = ["H"]
    cat_col = ["C"]
    title = "Avg GES message process time(ms)"
    chart_num = 4
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 5.sent Message ratio(%)
    data_col = ["I", "J", "K"]
    cat_col = ["C"]
    title = "sent Message ratio(%)"
    chart_num = 5
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 6.sent Messages
    data_col = ["L", "M", "N"]
    cat_col = ["C"]
    title = "sent Messages"
    chart_num = 6
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 07 RAC_GESSTAT]
    sheet = wb.create_sheet('rac_gesstat')
    wb.active = wb.sheetnames.index('rac_gesstat')
    sheet = wb.active

    sql = analyzeSQL.rac_gesstat
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)
    ## 그래프가 있을 경우 작성

    # [Step 08 RAC_GC_CR_SERV]
    sheet = wb.create_sheet('rac_gc_cr_serv')
    wb.active = wb.sheetnames.index('rac_gc_cr_serv')
    sheet = wb.active

    sql = analyzeSQL.rac_gc_cr_serv
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    ## 그래프가 있을 경우 작성
    max_cell_val = int(sheet.max_column) + 2

    # 0.Block Request
    data_col = ["D", "E", "F"]
    cat_col = ["C"]
    title = "Avg Message sent queue time(ms)"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 1.Request Type
    data_col = ["G", "H", "I"]
    cat_col = ["C"]
    title = "Request Type"
    chart_num = 1
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 2.Result Type
    data_col = ["J", "K", "L", "M", "N"]
    cat_col = ["C"]
    title = "Result Type"
    chart_num = 2
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 3.Fairness Down Convert(%)
    data_col = ["X"]
    cat_col = ["C"]
    title = "Fairness Down Convert(%)"
    chart_num = 3
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 09 RAC_GC_CUR_SRV]
    sheet = wb.create_sheet('rac_gc_cur_srv')
    wb.active = wb.sheetnames.index('rac_gc_cur_srv')
    sheet = wb.active

    sql = analyzeSQL.rac_gc_cur_srv
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    max_cell_val = int(sheet.max_column) + 2

    # 0.pin/flush/fusion write count
    data_col = ["I", "T"]
    cat_col = ["C"]
    title = "pin/flush/fusion write count"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 1. current block pin count
    data_col = ["D", "E", "F", "G", "H"]
    cat_col = ["C"]
    title = "current block pin count"
    chart_num = 1
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 2.current block pin(%)
    data_col = ["J", "K", "L", "M", "N"]
    cat_col = ["C"]
    title = "current block pin(%)"
    chart_num = 2
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 3.current block flush count
    data_col = ["O", "P", "Q", "R", "S"]
    cat_col = ["C"]
    title = "current block flush count"
    chart_num = 3
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 4.current block flush count(%)
    data_col = ["U", "V", "W", "X", "Y"]
    cat_col = ["C"]
    title = "current block flush count(%)"
    chart_num = 4
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 10 RAC_GC_INST_TRAN]
    sheet = wb.create_sheet('rac_gc_inst_tran')
    wb.active = wb.sheetnames.index('rac_gc_inst_tran')
    sheet = wb.active

    sql = analyzeSQL.rac_gc_inst_tran
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    max_cell_val = int(sheet.max_column) + 2

    # 0.cr block received
    data_col = ["E"]
    cat_col = ["C"]
    title = "cr block received"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 1.cr block transfer(%)
    data_col = ["F", "G", "H"]
    cat_col = ["C"]
    title = "cr block transfer(%)"
    chart_num = 1
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 2.current block received
    data_col = ["I"]
    cat_col = ["C"]
    title = "current block received"
    chart_num = 2
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 3.current block transfer(%)
    data_col = ["J", "K", "L"]
    cat_col = ["C"]
    title = "current block transfer(%)"
    chart_num = 3
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 4.avg cr time(ms)
    data_col = ["N", "O", "P"]
    cat_col = ["C"]
    title = "avg cr time(ms)"
    chart_num = 4
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 5.avg current time(ms)
    data_col = ["R", "S", "T"]
    cat_col = ["C"]
    title = "avg current time(ms)"
    chart_num = 5
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 11 RAC_PING]
    sheet = wb.create_sheet('rac_ping')
    wb.active = wb.sheetnames.index('rac_ping')
    sheet = wb.active

    sql = analyzeSQL.rac_ping
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    max_cell_val = int(sheet.max_column) + 2

    # 0.Ping Latency(500b)(ms)
    data_col = ["F"]
    cat_col = ["C"]
    title = "Ping Latency(500b)(ms)"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 1.Ping Latency(8k)(ms)
    data_col = ["H"]
    cat_col = ["C"]
    title = "Ping Latency(8k)(ms)"
    chart_num = 1
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 12 OS_STAT]
    sheet = wb.create_sheet('os_stat')
    wb.active = wb.sheetnames.index('os_stat')
    sheet = wb.active

    sql = analyzeSQL.os_stat
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    max_cell_val = int(sheet.max_column) + 2

    # 0.CPU Usage(s)
    data_col = ["D", "E", "F", "G"]
    cat_col = ["C"]
    title = "CPU Usage(s)"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 1.CPU usage(%)
    data_col = ["P", "Q", "R", "S"]
    cat_col = ["C"]
    title = "CPU usage(%)"
    chart_num = 1
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 2.CPU Count
    data_col = ["L", "M"]
    cat_col = ["C"]
    title = "CPU Count"
    chart_num = 2
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 3.Physical Memory(MB)
    data_col = ["K"]
    cat_col = ["C"]
    title = "Physical Memory(MB)"
    chart_num = 3
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 13 TIME_MODEL]
    sheet = wb.create_sheet('time_model')
    wb.active = wb.sheetnames.index('time_model')
    sheet = wb.active

    sql = analyzeSQL.time_model
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    max_cell_val = int(sheet.max_column) + 2

    # 0.DB Time vs background (s)
    data_col = ["D", "E", "F", "G"]
    cat_col = ["C"]
    title = "DB Time vs background (s)"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 1.DB Time stat(s)
    data_col = ["D", "E", "I", "K", "Q"]
    cat_col = ["C"]
    title = "DB Time stat(s)"
    chart_num = 1
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 2.sequence load elapsed time
    data_col = ["H"]
    cat_col = ["C"]
    title = "sequence load elapsed time"
    chart_num = 2
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 3.parse time elapsed
    data_col = ["I"]
    cat_col = ["C"]
    title = "parse time elapsed"
    chart_num = 3
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 4.hard parse elapsed time
    data_col = ["J"]
    cat_col = ["C"]
    title = "hard parse elapsed time"
    chart_num = 4
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 5.sql execute elapsed time
    data_col = ["K"]
    cat_col = ["C"]
    title = "sql execute elapsed time"
    chart_num = 5
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 6.connection management call elapsed time
    data_col = ["L"]
    cat_col = ["C"]
    title = "connection management call elapsed time"
    chart_num = 6
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 7.failed parse elapsed time
    data_col = ["M"]
    cat_col = ["C"]
    title = "failed parse elapsed time"
    chart_num = 7
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 8.failed parse (out of shared memory) elapsed time
    data_col = ["N"]
    cat_col = ["C"]
    title = "failed parse (out of shared memory) elapsed time"
    chart_num = 8
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 9.hard parse (sharing criteria) elapsed time
    data_col = ["O"]
    cat_col = ["C"]
    title = "hard parse (sharing criteria) elapsed time"
    chart_num = 9
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 10.hard parse (bind mismatch) elapsed time
    data_col = ["P"]
    cat_col = ["C"]
    title = "hard parse (bind mismatch) elapsed time"
    chart_num = 10
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 11.PL/SQL execution elapsed time
    data_col = ["Q"]
    cat_col = ["C"]
    title = "PL/SQL execution elapsed time"
    chart_num = 11
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 12.inbound PL/SQL rpc elapsed time
    data_col = ["R"]
    cat_col = ["C"]
    title = "inbound PL/SQL rpc elapsed time"
    chart_num = 12
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 13.PL/SQL compilation elapsed time
    data_col = ["S"]
    cat_col = ["C"]
    title = "PL/SQL compilation elapsed time"
    chart_num = 13
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 14.Java execution elapsed time
    data_col = ["T"]
    cat_col = ["C"]
    title = "Java execution elapsed time"
    chart_num = 14
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 14 WAIT5]
    sheet = wb.create_sheet('wait5')
    wb.active = wb.sheetnames.index('wait5')
    sheet = wb.active

    sql = analyzeSQL.wait5
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    # TOP Wait Event 5 컬럼명 수정
    sql = analyzeSQL.wait52
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    replace_column = cursor.fetchall()

    if not replace_column:
       pass
    else:
        j = 0
        for i in range(6, len(columns)):
            columns[i] = replace_column[j][0]
            j = j + 1

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    max_cell_val = int(sheet.max_column) + 2

    # 0.Top5 events-waits
    data_col = ["G", "H", "I", "J", "K"]
    cat_col = ["C"]
    title = "Top5 events-waits"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 1.Top5 events-wait time(s)
    data_col = ["L", "M", "N", "O", "P"]
    cat_col = ["C"]
    title = "Top5 events-wait time(s)"
    chart_num = 1
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 2.Top5 events-avg wait time(ms)
    data_col = ["Q", "R", "S", "T", "U"]
    cat_col = ["C"]
    title = "Top5 events-avg wait time(ms)"
    chart_num = 2
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)


    # [Step 15 WAIT20]
    sheet = wb.create_sheet('wait20')
    wb.active = wb.sheetnames.index('wait20')
    sheet = wb.active

    sql = analyzeSQL.wait20
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    # TOP Wait Event 20 컬럼명 수정
    sql = analyzeSQL.wait202
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    replace_column = cursor.fetchall()

    if not replace_column:
       pass
    else:
        j = 0
        for i in range(6, len(columns)):
            columns[i] = replace_column[j][0]
            j = j + 1

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    max_cell_val = int(sheet.max_column) + 2

    # 0.Top20 events-waits
    data_col = ["G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z"]
    cat_col = ["C"]
    title = "Top20 events-waits"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 1.Top20 events-wait time(s)
    data_col = ["AA", "AB", "AC", "AD", "AE", "AF", "AG", "AH", "AI", "AJ", "AK", "AL", "AM", "AN", "AO", "AP", "AQ", "AR", "AS", "AT"]
    cat_col = ["C"]
    title = "Top20 events-wait time(s)"
    chart_num = 1
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 2.Top20 events-avg wait time(ms)
    data_col = ["AU", "AV", "AW", "AX", "AY", "AZ", "BA", "BB", "BC", "BD", "BE", "BF", "BG", "BH", "BI", "BJ", "BK", "BL", "BM", "BM"]
    cat_col = ["C"]
    title = "Top20 events-avg wait time(ms)"
    chart_num = 2
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 16 WAIT_DISK]
    sheet = wb.create_sheet('wait_disk')
    wb.active = wb.sheetnames.index('wait_disk')
    sheet = wb.active

    sql = analyzeSQL.wait_disk
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    max_cell_val = int(sheet.max_column) + 2

    # 0.IO event(/sec)
    data_col = ["D", "E", "F", "G"]
    cat_col = ["C"]
    title = "IO event(/sec)"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 1.avg IO event time(ms)
    data_col = ["H", "I", "J", "K"]
    cat_col = ["C"]
    title = "avg IO event time(ms)"
    chart_num = 1
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 17 ENQUEUE5]
    sheet = wb.create_sheet('enqueue5')
    wb.active = wb.sheetnames.index('enqueue5')
    sheet = wb.active

    sql = analyzeSQL.enqueue5
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    # TOP ENQUEUE 5 컬럼명 수정
    sql = analyzeSQL.enqueue52
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    replace_column = cursor.fetchall()

    if not replace_column:
       pass
    elif len(replace_column) == 1:
        j = 0
        for i in range(3, len(columns) - 4):
            columns[i] = replace_column[j][0]
            j = j + 1
    elif len(replace_column) == 2:
        j = 0
        for i in range(3, len(columns) - 3):
            columns[i] = replace_column[j][0]
            j = j + 1
    elif len(replace_column) == 3:
        j = 0
        for i in range(3, len(columns) - 2):
            columns[i] = replace_column[j][0]
            j = j + 1
    elif len(replace_column) == 4:
        j = 0
        for i in range(3, len(columns) - 1):
            columns[i] = replace_column[j][0]
            j = j + 1
    else:
        j = 0
        for i in range(3, len(columns)):
            columns[i] = replace_column[j][0]
            j = j + 1

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    max_cell_val = int(sheet.max_column) + 2

    # 0.Top Enqueue Wait Time(s)
    data_col = ["D", "E", "F", "G", "H"]
    cat_col = ["C"]
    title = "Top Enqueue Wait Time(s)"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 18 ENQ_RAW_SQL]
    sheet = wb.create_sheet('enq_raw_sql')
    wb.active = wb.sheetnames.index('enq_raw_sql')
    sheet = wb.active

    sql = analyzeSQL.enq_raw_sql
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    # [Step 19 LATCH]
    sheet = wb.create_sheet('latch')
    wb.active = wb.sheetnames.index('latch')
    sheet = wb.active

    sql = analyzeSQL.latch
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    # TOP LATCH 5 컬럼명 수정
    sql = analyzeSQL.latch2
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    replace_column = cursor.fetchall()

    if not replace_column:
       pass
    else:
        j = 0
        for i in range(3, len(columns)):
            columns[i] = replace_column[j][0]
            j = j + 1

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    max_cell_val = int(sheet.max_column) + 2

    # 0.Top Latch Wait Time(ms/s)
    data_col = ["D", "E", "F", "G", "H"]
    cat_col = ["C"]
    title = "Top Latch Wait Time(ms/s)"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 1.Top Latch Miss(%)
    data_col = ["I", "J", "K", "L", "M"]
    cat_col = ["C"]
    title = "Top Latch Miss(%)"
    chart_num = 1
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 20 LATCH_SLEEP]
    sheet = wb.create_sheet('latch_sleep')
    wb.active = wb.sheetnames.index('latch_sleep')
    sheet = wb.active

    sql = analyzeSQL.latch_sleep
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    # [Step 21 LATCH_MISS]
    sheet = wb.create_sheet('latch_miss')
    wb.active = wb.sheetnames.index('latch_miss')
    sheet = wb.active

    sql = analyzeSQL.latch_miss
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    # [Step 22 ACTIVITY]
    sheet = wb.create_sheet('activity')
    wb.active = wb.sheetnames.index('activity')
    sheet = wb.active

    sql = analyzeSQL.activity
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    max_cell_val = int(sheet.max_column) + 2

    # 0.logon current
    data_col = ["D"]
    cat_col = ["C"]
    title = "logon current"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 1.open cursors current
    data_col = ["E"]
    cat_col = ["C"]
    title = "open cursors current"
    chart_num = 1
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 2.workarea memory allocated
    data_col = ["F"]
    cat_col = ["C"]
    title = "workarea memory allocated"
    chart_num = 2
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 3.session cursor cache count
    data_col = ["G"]
    cat_col = ["C"]
    title = "session cursor cache count"
    chart_num = 3
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 23 SQLSUM]
    sheet = wb.create_sheet('sqlsum')
    wb.active = wb.sheetnames.index('sqlsum')
    sheet = wb.active

    sql = analyzeSQL.sqlsum
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    max_cell_val = int(sheet.max_column) + 2

    # 0.sql execute count ratio(%)
    data_col = ["H", "I"]
    cat_col = ["C"]
    title = "sql execute count ratio(%)"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 1.sql memory ratio(%)
    data_col = ["J", "K"]
    cat_col = ["C"]
    title = "sql memory ratio(%)"
    chart_num = 1
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 24 TOP_SQL BY_QIO]
    sheet = wb.create_sheet('top_sql_by_qio')
    wb.active = wb.sheetnames.index('top_sql_by_qio')
    sheet = wb.active

    sql = analyzeSQL.top_sql_by_qio
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)
    df2 = df.sort_values(by="CPU_TIME", ascending=False).head(5)
    df3 = df.sort_values(by="ELAPSED_TIME", ascending=False).head(5)
    df4 = df.sort_values(by="BUFFER_GETS_DELTA", ascending=False).head(5)
    df5 = df.sort_values(by="DISK_READS_DELTA", ascending=False).head(5)
    df6 = df.sort_values(by="IOWAIT_DELTA", ascending=False).head(5)
    df7 = df.sort_values(by="PER_IO", ascending=False).head(5)

    df2 = df2.reset_index(drop=True)
    df3 = df3.reset_index(drop=True)
    df4 = df4.reset_index(drop=True)
    df5 = df5.reset_index(drop=True)
    df6 = df6.reset_index(drop=True)
    df7 = df7.reset_index(drop=True)

    df['TOP5_CPU_SQL'] = df2["SQL_ID"]
    df['TOP5_CPU_TIME'] = df2["CPU_TIME"]
    df['TOP5_ELAPS_SQL'] = df3["SQL_ID"]
    df['TOP5_ELAPS'] = df3["ELAPSED_TIME"]
    df['TOP5_BUFFER_SQL'] = df4["SQL_ID"]
    df['TOP5_BUFFER'] = df4["BUFFER_GETS_DELTA"]
    df['TOP5_DISK_SQL'] = df5["SQL_ID"]
    df['TOP5_DISK'] = df5["DISK_READS_DELTA"]
    df['TOP5_IOWAIT_SQL'] = df6["SQL_ID"]
    df['TOP5_IOWAIT'] = df6["IOWAIT_DELTA"]
    df['TOP5_PERIO_SQL'] = df7["SQL_ID"]
    df['TOP5_PERIO'] = df7["PER_IO"]

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    max_cell_val = int(sheet.max_column) + 2

    # 0.Top SQL by CPU time
    data_col = ["V"]
    cat_col = ["U"]
    title = "Top SQL by CPU time"
    chart_num = 0
    top = 5
    PieChart(data_col, cat_col, title, sheet, max_cell_val, chart_num, top)

    # 1.Top SQL by Elapsed time
    data_col = ["X"]
    cat_col = ["W"]
    title = "Top SQL by Elapsed time"
    chart_num = 1
    top = 5
    PieChart(data_col, cat_col, title, sheet, max_cell_val, chart_num, top)

    # 2.Top SQL by Buffer gets
    data_col = ["Z"]
    cat_col = ["Y"]
    title = "Top SQL by Buffer gets"
    chart_num = 2
    top = 5
    PieChart(data_col, cat_col, title, sheet, max_cell_val, chart_num, top)

    # 3.Top SQL by Disk Reads
    data_col = ["AB"]
    cat_col = ["AA"]
    title = "Top SQL by Disk Reads"
    chart_num = 3
    top = 5
    PieChart(data_col, cat_col, title, sheet, max_cell_val, chart_num, top)

    # 4.Top SQL by IOWAIT
    data_col = ["AD"]
    cat_col = ["AC"]
    title = "Top SQL by IOWAIT"
    chart_num = 4
    top = 5
    PieChart(data_col, cat_col, title, sheet, max_cell_val, chart_num, top)

    # 5.Top SQL by Per IO
    data_col = ["AF"]
    cat_col = ["AE"]
    title = "Top SQL by Per IO"
    chart_num = 5
    top = 5
    PieChart(data_col, cat_col, title, sheet, max_cell_val, chart_num, top)

    # [Step 25 SQL_ELAPS]
    sheet = wb.create_sheet('sql_elaps')
    wb.active = wb.sheetnames.index('sql_elaps')
    sheet = wb.active

    sql = analyzeSQL.sql_elaps
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    # TOP SQL ELAPS 5 컬럼명 수정
    sql = analyzeSQL.sql_elaps2
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    replace_column = cursor.fetchall()

    if not replace_column:
       pass
    else:
        j = 0
        for i in range(3, len(columns)-5):
            columns[i] = replace_column[j][0]
            j = j + 1

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    max_cell_val = int(sheet.max_column) + 2

    # 0.Top Elapsed sql-elapsed time(s)
    data_col = ["D", "E", "F", "G", "H"]
    cat_col = ["C"]
    title = "Top Elapsed sql-elapsed time(s)"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 1.Top Elapsed sql-cpu time(s)
    data_col = ["I", "J", "K", "L", "M"]
    cat_col = ["C"]
    title = "Top Elapsed sql-cpu time(s)"
    chart_num = 1
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 2.Top Elapsed sql-executions
    data_col = ["N", "O", "P", "Q", "R"]
    cat_col = ["C"]
    title = "Top Elapsed sql-executions"
    chart_num = 2
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 3.Top Elapsed sql-elapsed time/execute(s)
    data_col = ["S", "T", "U", "V", "W"]
    cat_col = ["C"]
    title = "Top Elapsed sql-elapsed time/execute(s)"
    chart_num = 3
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 4.Top Elapsed sql-elapsed/db time(%)
    data_col = ["X", "Y", "Z", "AA", "AB"]
    cat_col = ["C"]
    title = "Top Elapsed sql-elapsed/db time(%)"
    chart_num = 4
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 26 SQL_ELAPS_SQL]
    sheet = wb.create_sheet('sql_elaps_sql')
    wb.active = wb.sheetnames.index('sql_elaps_sql')
    sheet = wb.active

    sql = analyzeSQL.sql_elaps_sql
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    # [Step 27 SQL_ELAPS_RAW]
    sheet = wb.create_sheet('sql_elaps_raw')
    wb.active = wb.sheetnames.index('sql_elaps_raw')
    sheet = wb.active

    sql = analyzeSQL.sql_elaps_raw
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    # [Step 28 SQL_CPU]
    sheet = wb.create_sheet('sql_cpu')
    wb.active = wb.sheetnames.index('sql_cpu')
    sheet = wb.active

    sql = analyzeSQL.sql_cpu
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    # TOP SQL CPU 5 컬럼명 수정
    sql = analyzeSQL.sql_cpu2
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    replace_column = cursor.fetchall()

    if not replace_column:
       pass
    else:
        j = 0
        for i in range(3, len(columns)-5):
            columns[i] = replace_column[j][0]
            j = j + 1

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    max_cell_val = int(sheet.max_column) + 2

    # 0.Top CPU sql-cpu time(s)
    data_col = ["D", "E", "F", "G", "H"]
    cat_col = ["C"]
    title = "Top CPU sql-cpu time(s)"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 1.Top CPU sql-elapsed time(s)
    data_col = ["I", "J", "K", "L", "M"]
    cat_col = ["C"]
    title = "Top CPU sql-elapsed time(s)"
    chart_num = 1
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 2.Top CPU sql-executions
    data_col = ["N", "O", "P", "Q", "R"]
    cat_col = ["C"]
    title = "Top CPU sql-executions"
    chart_num = 2
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 3.Top CPU sql-cpu time/execute(s)
    data_col = ["S", "T", "U", "V", "W"]
    cat_col = ["C"]
    title = "Top CPU sql-cpu time/execute(s)"
    chart_num = 3
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 4.Top CPU elapsed time/db time(%)
    data_col = ["X", "Y", "Z", "AA", "AB"]
    cat_col = ["C"]
    title = "Top CPU elapsed time/db time(%)"
    chart_num = 4
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 29 SQL_CPU_SQL]
    sheet = wb.create_sheet('sql_cpu_sql')
    wb.active = wb.sheetnames.index('sql_cpu_sql')
    sheet = wb.active

    sql = analyzeSQL.sql_cpu_sql
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    # [Step 30 SQL_CPU_RAW]
    sheet = wb.create_sheet('sql_cpu_raw')
    wb.active = wb.sheetnames.index('sql_cpu_raw')
    sheet = wb.active

    sql = analyzeSQL.sql_cpu_raw
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    # [Step 31 SQL_GET]
    sheet = wb.create_sheet('sql_get')
    wb.active = wb.sheetnames.index('sql_get')
    sheet = wb.active

    sql = analyzeSQL.sql_get
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    # TOP SQL GET 5 컬럼명 수정
    sql = analyzeSQL.sql_get2
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    replace_column = cursor.fetchall()

    if not replace_column:
       pass
    else:
        j = 0
        for i in range(3, len(columns)-5):
            columns[i] = replace_column[j][0]
            j = j + 1

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    max_cell_val = int(sheet.max_column) + 2

    # 0.Top Bufferget sql-buffer gets
    data_col = ["D", "E", "F", "G", "H"]
    cat_col = ["C"]
    title = "Top Bufferget sql-buffer gets"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 1.Top Bufferget sql-cpu time(s)
    data_col = ["I", "J", "K", "L", "M"]
    cat_col = ["C"]
    title = "Top Bufferget sql-cpu time(s)"
    chart_num = 1
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 2.Top Bufferget sql-elapsed time(s)
    data_col = ["N", "O", "P", "Q", "R"]
    cat_col = ["C"]
    title = "Top Bufferget sql-elapsed time(s)"
    chart_num = 2
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 3.Top Bufferget sql-executions
    data_col = ["S", "T", "U", "V", "W"]
    cat_col = ["C"]
    title = "Top Bufferget sql-executions"
    chart_num = 3
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 4.Top Bufferget sql-buffer get/execute
    data_col = ["X", "Y", "Z", "AA", "AB"]
    cat_col = ["C"]
    title = "Top Bufferget sql-buffer get/execute"
    chart_num = 4
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 5.Top Bufferget sql-buffer get(%)
    data_col = ["AC", "AD", "AE", "AF", "AG"]
    cat_col = ["C"]
    title = "Top Bufferget sql-buffer get(%)"
    chart_num = 5
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 32 SQL_GET_SQL]
    sheet = wb.create_sheet('sql_get_sql')
    wb.active = wb.sheetnames.index('sql_get_sql')
    sheet = wb.active

    sql = analyzeSQL.sql_get_sql
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    # [Step 33 SQL_GET_RAW]
    sheet = wb.create_sheet('sql_get_raw')
    wb.active = wb.sheetnames.index('sql_get_raw')
    sheet = wb.active

    sql = analyzeSQL.sql_get_raw
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    # [Step 34 SQL_IO]
    sheet = wb.create_sheet('sql_io')
    wb.active = wb.sheetnames.index('sql_io')
    sheet = wb.active

    sql = analyzeSQL.sql_io
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    # TOP SQL IO 5 컬럼명 수정
    sql = analyzeSQL.sql_io2
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    replace_column = cursor.fetchall()

    if not replace_column:
       pass
    else:
        j = 0
        for i in range(3, len(columns)-5):
            columns[i] = replace_column[j][0]
            j = j + 1

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    max_cell_val = int(sheet.max_column) + 2

    # 0.Top Disk reads sql-disk reads
    data_col = ["D", "E", "F", "G", "H"]
    cat_col = ["C"]
    title = "Top Disk reads sql-disk reads"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 1.Top Disk reads sql-cpu time(s)
    data_col = ["I", "J", "K", "L", "M"]
    cat_col = ["C"]
    title = "Top Disk reads sql-cpu time(s)"
    chart_num = 1
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 2.Top Disk reads sql-elapsed time(s)
    data_col = ["N", "O", "P", "Q", "R"]
    cat_col = ["C"]
    title = "Top Disk reads sql-elapsed time(s)"
    chart_num = 2
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 3.Top Disk reads sql-executions
    data_col = ["S", "T", "U", "V", "W"]
    cat_col = ["C"]
    title = "Top Disk reads sql-executions"
    chart_num = 3
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 4.Top Disk reads sql-buffer get/execute
    data_col = ["X", "Y", "Z", "AA", "AB"]
    cat_col = ["C"]
    title = "Top Disk reads sql-buffer get/execute"
    chart_num = 4
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 5.Top Disk reads sql-disk reads(%)
    data_col = ["AC", "AD", "AE", "AF", "AG"]
    cat_col = ["C"]
    title = "Top Disk reads sql-disk reads(%)"
    chart_num = 5
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 35 SQL_IO_SQL]
    sheet = wb.create_sheet('sql_io_sql')
    wb.active = wb.sheetnames.index('sql_io_sql')
    sheet = wb.active

    sql = analyzeSQL.sql_io_sql
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    # [Step 36 SQL_IO_RAW]
    sheet = wb.create_sheet('sql_io_raw')
    wb.active = wb.sheetnames.index('sql_io_raw')
    sheet = wb.active

    sql = analyzeSQL.sql_io_raw
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    # [Step 37 SQL_EXEC]
    sheet = wb.create_sheet('sql_exec')
    wb.active = wb.sheetnames.index('sql_exec')
    sheet = wb.active

    sql = analyzeSQL.sql_exec
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    # TOP SQL EXEC 5 컬럼명 수정
    sql = analyzeSQL.sql_exec2
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    replace_column = cursor.fetchall()

    if not replace_column:
       pass
    else:
        j = 0
        for i in range(3, len(columns)-5):
            columns[i] = replace_column[j][0]
            j = j + 1

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)
    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    max_cell_val = int(sheet.max_column) + 2

    # 0.Top Execution sql-executions
    data_col = ["D", "E", "F", "G", "H"]
    cat_col = ["C"]
    title = "Top Execution sql-executions"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 1.Top Execution sql-rows processed
    data_col = ["I", "J", "K", "L", "M"]
    cat_col = ["C"]
    title = "Top Execution sql-rows processed"
    chart_num = 1
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 2.Top Execution sql-row_processed/execute
    data_col = ["N", "O", "P", "Q", "R"]
    cat_col = ["C"]
    title = "Top Execution sql-row_processed/execute"
    chart_num = 2
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 3.Top Execution sql-cpu time/execute(s)
    data_col = ["S", "T", "U", "V", "W"]
    cat_col = ["C"]
    title = "Top Execution sql-cpu time/execute(s)"
    chart_num = 3
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 4.Top Execution sql-elapsed time/execute(s)
    data_col = ["X", "Y", "Z", "AA", "AB"]
    cat_col = ["C"]
    title = "Top Execution sql-elapsed time/execute(s)"
    chart_num = 4
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 38 SQL_EXEC_SQL]
    sheet = wb.create_sheet('sql_exec_sql')
    wb.active = wb.sheetnames.index('sql_exec_sql')
    sheet = wb.active

    sql = analyzeSQL.sql_exec_sql
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    # [Step 39 SQL_EXEC_RAW]
    sheet = wb.create_sheet('sql_exec_raw')
    wb.active = wb.sheetnames.index('sql_exec_raw')
    sheet = wb.active

    sql = analyzeSQL.sql_exec_raw
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    # [Step 40 SQL_CLWAIT]
    sheet = wb.create_sheet('sql_clwait')
    wb.active = wb.sheetnames.index('sql_clwait')
    sheet = wb.active

    sql = analyzeSQL.sql_clwait
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    # TOP SQL CLWAIT 5 컬럼명 수정
    sql = analyzeSQL.sql_clwait2
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    replace_column = cursor.fetchall()

    if not replace_column:
       pass
    else:
        j = 0
        for i in range(3, len(columns)-5):
            columns[i] = replace_column[j][0]
            j = j + 1

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)
    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    max_cell_val = int(sheet.max_column) + 2

    # 0.Top Cluster Wait sql-cluster wait time(s)
    data_col = ["D", "E", "F", "G", "H"]
    cat_col = ["C"]
    title = "Top Cluster Wait sql-cluster wait time(s)"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 1.Top Cluster Wait sql-cpu time(s)
    data_col = ["I", "J", "K", "L", "M"]
    cat_col = ["C"]
    title = "Top Cluster Wait sql-cpu time(s)"
    chart_num = 1
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 2.Top Cluster Wait sql-elpased time(s)
    data_col = ["N", "O", "P", "Q", "R"]
    cat_col = ["C"]
    title = "Top Cluster Wait sql-elpased time(s)"
    chart_num = 2
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 3.Top Cluster Wait sql-executions
    data_col = ["S", "T", "U", "V", "W"]
    cat_col = ["C"]
    title = "Top Cluster Wait sql-executions"
    chart_num = 3
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 4.Top Cluster Wait sql-cluster wait time/elapsed time(%)
    data_col = ["X", "Y", "Z", "AA", "AB"]
    cat_col = ["C"]
    title = "Top Cluster Wait sql-cluster wait time/elapsed time(%)"
    chart_num = 4
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 41 SQL_CLWAIT_SQL]
    sheet = wb.create_sheet('sql_clwait_sql')
    wb.active = wb.sheetnames.index('sql_clwait_sql')
    sheet = wb.active

    sql = analyzeSQL.sql_clwait_sql
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)
    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    # [Step 42 SQL_CLWAIT_RAW]
    sheet = wb.create_sheet('sql_clwait_raw')
    wb.active = wb.sheetnames.index('sql_clwait_raw')
    sheet = wb.active

    sql = analyzeSQL.sql_clwait_raw
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)
    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    # [Step 43 SEG_LOGICAL]
    sheet = wb.create_sheet('seg_logical')
    wb.active = wb.sheetnames.index('seg_logical')
    sheet = wb.active

    sql = analyzeSQL.seg_logical
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    # TOP SEG_LOGICAL 5 컬럼명 수정
    sql = analyzeSQL.seg_logical2
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    replace_column = cursor.fetchall()

    if not replace_column:
       pass
    else:
        j = 0
        for i in range(3, len(columns)):
            columns[i] = replace_column[j][0]
            j = j + 1

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)
    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)
    max_cell_val = int(sheet.max_column) + 2

    # 0.Top Logical Reads Segment-logical reads
    data_col = ["D", "E", "F", "G", "H"]
    cat_col = ["C"]
    title = "Top Logical Reads Segment-logical reads"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 1.Top Logical Reads Segment-logical reads(%)
    data_col = ["I", "J", "K", "L", "M"]
    cat_col = ["C"]
    title = "Top Logical Reads Segment-logical reads(%)"
    chart_num = 1
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 2.Top Logical Reads Segment-logical reads(/sec)
    data_col = ["N", "O", "P", "Q", "R"]
    cat_col = ["C"]
    title = "Top Logical Reads Segment-logical reads(/sec)"
    chart_num = 2
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 44 SEG_PHYSICAL]
    sheet = wb.create_sheet('seg_physical')
    wb.active = wb.sheetnames.index('seg_physical')
    sheet = wb.active

    sql = analyzeSQL.seg_physical
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    # TOP SEG_PHYSICAL 5 컬럼명 수정
    sql = analyzeSQL.seg_physical2
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    replace_column = cursor.fetchall()

    if not replace_column:
       pass
    else:
        j = 0
        for i in range(3, len(columns)):
            columns[i] = replace_column[j][0]
            j = j + 1

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)
    max_cell_val = int(sheet.max_column) + 2

    # 0.Top Physical Reads Segment-physical reads
    data_col = ["D", "E", "F", "G", "H"]
    cat_col = ["C"]
    title = "Top Physical Reads Segment-physical reads"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 1.Top Physical Reads Segment-physical reads(%)
    data_col = ["I", "J", "K", "L", "M"]
    cat_col = ["C"]
    title = "Top Physical Reads Segment-physical reads(%)"
    chart_num = 1
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 2.Top Physical Reads Segment-physical reads(/sec)
    data_col = ["N", "O", "P", "Q", "R"]
    cat_col = ["C"]
    title = "Top Physical Reads Segment-physical reads(/sec)"
    chart_num = 2
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 45 SEG_BUFFBUSY]
    sheet = wb.create_sheet('seg_buffbusy')
    wb.active = wb.sheetnames.index('seg_buffbusy')
    sheet = wb.active

    sql = analyzeSQL.seg_buffbusy
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    # TOP SEG_BUFFBUSY 5 컬럼명 수정
    sql = analyzeSQL.seg_buffbusy2
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    replace_column = cursor.fetchall()

    if not replace_column:
       pass
    else:
        j = 0
        for i in range(3, len(columns)):
            columns[i] = replace_column[j][0]
            j = j + 1

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)
    max_cell_val = int(sheet.max_column) + 2

    # 0.Top Buff. busy wait Segment-buff. busy waits
    data_col = ["D", "E", "F", "G", "H"]
    cat_col = ["C"]
    title = "Top Buff. busy wait Segment-buff. busy waits"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 1.Top Buff. busy wait Segment-buff. busy waits(%)
    data_col = ["I", "J", "K", "L", "M"]
    cat_col = ["C"]
    title = "Top Buff. busy wait Segment-buff. busy waits(%)"
    chart_num = 1
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 2.Top Buff. busy wait Segment-buff. busy waits(/sec)
    data_col = ["N", "O", "P", "Q", "R"]
    cat_col = ["C"]
    title = "Top Buff. busy wait Segment-buff. busy waits(/sec)"
    chart_num = 2
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 46 SEG_ROWLOCK]
    sheet = wb.create_sheet('seg_rowlock')
    wb.active = wb.sheetnames.index('seg_rowlock')
    sheet = wb.active

    sql = analyzeSQL.seg_rowlock
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    # TOP SEG_ROWLOCK 5 컬럼명 수정
    sql = analyzeSQL.seg_rowlock2
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    replace_column = cursor.fetchall()

    if not replace_column:
       pass
    else:
        j = 0
        for i in range(3, len(columns)):
            columns[i] = replace_column[j][0]
            j = j + 1

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)
    max_cell_val = int(sheet.max_column) + 2

    # 0.Top Row lock wait Segment-row lock waits
    data_col = ["D", "E", "F", "G", "H"]
    cat_col = ["C"]
    title = "Top Row lock wait Segment-row lock waits"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 1.Top Row lock wait Segment-row lock waits(%)
    data_col = ["I", "J", "K", "L", "M"]
    cat_col = ["C"]
    title = "Top Row lock wait Segment-row lock waits(%)"
    chart_num = 1
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 2.Top Row lock wait Segment-row lock waits(/sec)
    data_col = ["N", "O", "P", "Q", "R"]
    cat_col = ["C"]
    title = "Top Row lock wait Segment-row lock waits(/sec)"
    chart_num = 2
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 47 SEG_itl]
    sheet = wb.create_sheet('seg_itl')
    wb.active = wb.sheetnames.index('seg_itl')
    sheet = wb.active

    sql = analyzeSQL.seg_itl
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    # TOP SEG_itl 5 컬럼명 수정
    sql = analyzeSQL.seg_itl2
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    replace_column = cursor.fetchall()

    if not replace_column:
       pass
    else:
        j = 0
        for i in range(3, len(columns)):
            columns[i] = replace_column[j][0]
            j = j + 1

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)
    max_cell_val = int(sheet.max_column) + 2

    # 0.Top itl wait Segment-itl waits
    data_col = ["D", "E", "F", "G", "H"]
    cat_col = ["C"]
    title = "Top itl wait Segment-itl waits"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 1.Top itl wait Segment-itl waits(%)
    data_col = ["I", "J", "K", "L", "M"]
    cat_col = ["C"]
    title = "Top itl wait Segment-itl waits(%)"
    chart_num = 1
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 2.Top itl wait Segment-itl waits(/sec)
    data_col = ["N", "O", "P", "Q", "R"]
    cat_col = ["C"]
    title = "Top itl wait Segment-itl waits(/sec)"
    chart_num = 2
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 48 SEG_SPACEUSED]
    sheet = wb.create_sheet('seg_spaceused')
    wb.active = wb.sheetnames.index('seg_spaceused')
    sheet = wb.active

    sql = analyzeSQL.seg_spaceused
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    # TOP SEG_SPACEUSED 5 컬럼명 수정
    sql = analyzeSQL.seg_spaceused2
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    replace_column = cursor.fetchall()

    if not replace_column:
       pass
    else:
        j = 0
        for i in range(3, len(columns)):
            columns[i] = replace_column[j][0]
            j = j + 1

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)
    max_cell_val = int(sheet.max_column) + 2

    # 0.Top space used Segment-space used
    data_col = ["D", "E", "F", "G", "H"]
    cat_col = ["C"]
    title = "Top space used Segment-space used"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 1.Top space used Segment-space used(%)
    data_col = ["I", "J", "K", "L", "M"]
    cat_col = ["C"]
    title = "Top space used Segment-space used(%)"
    chart_num = 1
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 2.Top space used Segment-space used(/sec)
    data_col = ["N", "O", "P", "Q", "R"]
    cat_col = ["C"]
    title = "Top space used Segment-space used(/sec)"
    chart_num = 2
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 49 SEG_GCCRSERV]
    sheet = wb.create_sheet('seg_gccrserv')
    wb.active = wb.sheetnames.index('seg_gccrserv')
    sheet = wb.active

    sql = analyzeSQL.seg_gccrserv
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    # TOP SEG_GCCRSERV 5 컬럼명 수정
    sql = analyzeSQL.seg_gccrserv2
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    replace_column = cursor.fetchall()

    if not replace_column:
       pass
    else:
        j = 0
        for i in range(3, len(columns)):
            columns[i] = replace_column[j][0]
            j = j + 1

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)
    max_cell_val = int(sheet.max_column) + 2

    # 0.Top gc cr blaock served Segment-cr block served
    data_col = ["D", "E", "F", "G", "H"]
    cat_col = ["C"]
    title = "Top gc cr blaock served Segment-cr block served"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 1.Top gc cr block served Segment-cr block served(%)
    data_col = ["I", "J", "K", "L", "M"]
    cat_col = ["C"]
    title = "Top gc cr block served Segment-cr block served(%)"
    chart_num = 1
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 2.Top gc cr block served Segment-cr block served(/sec)
    data_col = ["N", "O", "P", "Q", "R"]
    cat_col = ["C"]
    title = "Top gc cr block served Segment-cr block served(/sec)"
    chart_num = 2
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 50 SEG_GCCUSERV]
    sheet = wb.create_sheet('seg_gccuserv')
    wb.active = wb.sheetnames.index('seg_gccuserv')
    sheet = wb.active

    sql = analyzeSQL.seg_gccuserv
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    # TOP SEG_GCCUSERV 5 컬럼명 수정
    sql = analyzeSQL.seg_gccuserv2
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    replace_column = cursor.fetchall()

    if not replace_column:
       pass
    else:
        j = 0
        for i in range(3, len(columns)):
            columns[i] = replace_column[j][0]
            j = j + 1

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)
    max_cell_val = int(sheet.max_column) + 2

    # 0.Top gc current block served Segment-current block served
    data_col = ["D", "E", "F", "G", "H"]
    cat_col = ["C"]
    title = "Top gc current block served Segment-current block served"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 1.Top gc current block served Segment-current block served(%)
    data_col = ["I", "J", "K", "L", "M"]
    cat_col = ["C"]
    title = "Top gc current block served Segment-current block served(%)"
    chart_num = 1
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 2.Top gc current block served Segment-current block served(/sec)
    data_col = ["N", "O", "P", "Q", "R"]
    cat_col = ["C"]
    title = "Top gc current block served Segment-current block served(/sec)"
    chart_num = 2
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 51 SEG_GCCRRCVD]
    sheet = wb.create_sheet('seg_gccrrcvd')
    wb.active = wb.sheetnames.index('seg_gccrrcvd')
    sheet = wb.active

    sql = analyzeSQL.seg_gccrrcvd
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    # TOP SEG_GCCRRCVD 5 컬럼명 수정
    sql = analyzeSQL.seg_gccrrcvd2
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    replace_column = cursor.fetchall()

    if not replace_column:
       pass
    else:
        j = 0
        for i in range(3, len(columns)):
            columns[i] = replace_column[j][0]
            j = j + 1

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)
    max_cell_val = int(sheet.max_column) + 2

    # 0.Top gc cr block received Segment-cr block received
    data_col = ["D", "E", "F", "G", "H"]
    cat_col = ["C"]
    title = "Top gc cr block received Segment-cr block received"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 1.Top gc cr block received Segment-cr block received(%)
    data_col = ["I", "J", "K", "L", "M"]
    cat_col = ["C"]
    title = "Top gc cr block received Segment-cr block received(%)"
    chart_num = 1
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 2.Top gc current block received Segment-current block received(/sec)
    data_col = ["N", "O", "P", "Q", "R"]
    cat_col = ["C"]
    title = "Top gc current block received Segment-current block received(/sec)"
    chart_num = 2
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 52 SEG_GCCURCVD]
    sheet = wb.create_sheet('seg_gccurcvd')
    wb.active = wb.sheetnames.index('seg_gccurcvd')
    sheet = wb.active

    sql = analyzeSQL.seg_gccurcvd
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    # TOP SEG_GCCURCVD 5 컬럼명 수정
    sql = analyzeSQL.seg_gccurcvd2
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    replace_column = cursor.fetchall()

    if not replace_column:
       pass
    else:
        j = 0
        for i in range(3, len(columns)):
            columns[i] = replace_column[j][0]
            j = j + 1

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)
    max_cell_val = int(sheet.max_column) + 2

    # 0.Top gc current block received Segment-current block received
    data_col = ["D", "E", "F", "G", "H"]
    cat_col = ["C"]
    title = "Top gc current block received Segment-current block received"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 1.Top gc current block received Segment-current block received(%)
    data_col = ["I", "J", "K", "L", "M"]
    cat_col = ["C"]
    title = "Top gc current block received Segment-current block received(%)"
    chart_num = 1
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 2.Top gc current block received Segment-current block received(/sec)
    data_col = ["N", "O", "P", "Q", "R"]
    cat_col = ["C"]
    title = "Top gc current block received Segment-current block received(/sec)"
    chart_num = 2
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 53 SEG_GCBUFFBUSY]
    sheet = wb.create_sheet('seg_gcbuffbusy')
    wb.active = wb.sheetnames.index('seg_gcbuffbusy')
    sheet = wb.active

    sql = analyzeSQL.seg_gcbuffbusy
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    # TOP SEG_GCBUFFBUSY 5 컬럼명 수정
    sql = analyzeSQL.seg_gcbuffbusy2
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    replace_column = cursor.fetchall()

    if not replace_column:
       pass
    else:
        j = 0
        for i in range(3, len(columns)):
            columns[i] = replace_column[j][0]
            j = j + 1

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)
    max_cell_val = int(sheet.max_column) + 2

    # 0.Top gc buffer busy wait Segment-gc buffer busy waits
    data_col = ["D", "E", "F", "G", "H"]
    cat_col = ["C"]
    title = "Top gc buffer busy wait Segment-gc buffer busy waits"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 1.Top gc buffer busy wait Segment-gc buffer busy waits(%)
    data_col = ["I", "J", "K", "L", "M"]
    cat_col = ["C"]
    title = "Top gc buffer busy wait Segment-gc buffer busy waits(%)"
    chart_num = 1
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 2.Top gc buffer busy wait Segment-gc buffer busy waits(/sec)
    data_col = ["N", "O", "P", "Q", "R"]
    cat_col = ["C"]
    title = "Top gc buffer busy wait Segment-gc buffer busy waits(/sec)"
    chart_num = 2
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 54 BLK_WAITSTAT_MS]
    sheet = wb.create_sheet('blk_waitstat_ms')
    wb.active = wb.sheetnames.index('blk_waitstat_ms')
    sheet = wb.active

    sql = analyzeSQL.blk_waitstat_ms
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    # [Step 55 TBSSTAT_TOTAL]
    sheet = wb.create_sheet('tbsstat_total')
    wb.active = wb.sheetnames.index('tbsstat_total')
    sheet = wb.active

    sql = analyzeSQL.tbsstat_total
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    # [Step 56 FILESTAT_TOTAL]
    sheet = wb.create_sheet('filestat_total')
    wb.active = wb.sheetnames.index('filestat_total')
    sheet = wb.active

    sql = analyzeSQL.filestat_total
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    # [Step 57 TBS_BLKRW]
    sheet = wb.create_sheet('tbs_blkrw')
    wb.active = wb.sheetnames.index('tbs_blkrw')
    sheet = wb.active

    sql = analyzeSQL.tbs_blkrw
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    # TOP TBS_BLKRW 5 컬럼명 수정
    sql = analyzeSQL.tbs_blkrw2
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    replace_column = cursor.fetchall()
    
    if not replace_column:
       pass
    else:
        j = 0
        for i in range(3, len(columns)):
            columns[i] = replace_column[j][0]
            j = j + 1

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)
    max_cell_val = int(sheet.max_column) + 2

    # 0.tablespace block read+write(/sec)
    data_col = ["D", "E", "F", "G", "H"]
    cat_col = ["C"]
    title = "tablespace block read+write(/sec)"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 58 FILES_BLKRW]
    sheet = wb.create_sheet('files_blkrw')
    wb.active = wb.sheetnames.index('files_blkrw')
    sheet = wb.active

    sql = analyzeSQL.files_blkrw
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    # TOP FILES_BLKRW 5 컬럼명 수정
    sql = analyzeSQL.files_blkrw2
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    replace_column = cursor.fetchall()

    if not replace_column:
       pass
    else:
        j = 0
        for i in range(3, len(columns)):
            columns[i] = replace_column[j][0]
            j = j + 1

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)
    max_cell_val = int(sheet.max_column) + 2

    # 0.datafile block read+write(/sec)
    data_col = ["D", "E", "F", "G", "H"]
    cat_col = ["C"]
    title = "datafile block read+write(/sec)"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 59 FILES_AVG_BLKS_READ]
    sheet = wb.create_sheet('files_avg_blks_read')
    wb.active = wb.sheetnames.index('files_avg_blks_read')
    sheet = wb.active

    sql = analyzeSQL.files_avg_blks_read
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    # TOP FILES_AVG_BLKS_READ 5 컬럼명 수정
    sql = analyzeSQL.files_avg_blks_read2
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    replace_column = cursor.fetchall()

    if not replace_column:
       pass
    else:
        j = 0
        for i in range(3, len(columns)):
            columns[i] = replace_column[j][0]
            j = j + 1

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)
    max_cell_val = int(sheet.max_column) + 2

    # 0.avg blocks/read
    data_col = ["D", "E", "F", "G", "H"]
    cat_col = ["C"]
    title = "avg blocks/read"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 60 FILES_AVG_TIME_READ]
    sheet = wb.create_sheet('files_avg_time_read')
    wb.active = wb.sheetnames.index('files_avg_time_read')
    sheet = wb.active

    sql = analyzeSQL.files_avg_time_read
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    # TOP FILES_AVG_TIME_READ 5 컬럼명 수정
    sql = analyzeSQL.files_avg_time_read2
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    replace_column = cursor.fetchall()

    if not replace_column:
       pass
    else:
        j = 0
        for i in range(3, len(columns)):
            columns[i] = replace_column[j][0]
            j = j + 1

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)
    max_cell_val = int(sheet.max_column) + 2

    # 0.avg readtime/phyblkread(ms)
    data_col = ["D", "E", "F", "G", "H"]
    cat_col = ["C"]
    title = "avg readtime/phyblkread(ms)"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 61 FILES_WAIT_TIME]
    sheet = wb.create_sheet('files_wait_time')
    wb.active = wb.sheetnames.index('files_wait_time')
    sheet = wb.active

    sql = analyzeSQL.files_wait_time
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    # TOP FILES_WAIT_TIME 5 컬럼명 수정
    sql = analyzeSQL.files_wait_time2
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    replace_column = cursor.fetchall()

    if not replace_column:
       pass
    else:
        j = 0
        for i in range(3, len(columns)):
            columns[i] = replace_column[j][0]
            j = j + 1

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)
    max_cell_val = int(sheet.max_column) + 2

    # 0.avg wait time(ms)
    data_col = ["D", "E", "F", "G", "H"]
    cat_col = ["C"]
    title = "avg wait time(ms)"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 62 V_PGA]
    sheet = wb.create_sheet('v_pga')
    wb.active = wb.sheetnames.index('v_pga')
    sheet = wb.active

    sql = analyzeSQL.v_pga
    cursor.execute(sql)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    # [Step 63 PGA]
    sheet = wb.create_sheet('pga')
    wb.active = wb.sheetnames.index('pga')
    sheet = wb.active

    sql = analyzeSQL.pga
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    max_cell_val = int(sheet.max_column) + 2

    # 0.pga size (MB)
    data_col = ["D", "E", "F", "G", "H", "I"]
    cat_col = ["C"]
    title = "pga size (MB)"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 64 PGA_CACHEHIT]
    sheet = wb.create_sheet('pga_cachehit')
    wb.active = wb.sheetnames.index('pga_cachehit')
    sheet = wb.active

    sql = analyzeSQL.pga_cachehit
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)
    max_cell_val = int(sheet.max_column) + 2

    # 0.cache hit(%)
    data_col = ["D"]
    cat_col = ["C"]
    title = "cache hit(%)"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 65 WORKAREA]
    sheet = wb.create_sheet('workarea')
    wb.active = wb.sheetnames.index('workarea')
    sheet = wb.active

    sql = analyzeSQL.workarea
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    # [Step 66 OPTIMAL]
    sheet = wb.create_sheet('optimal')
    wb.active = wb.sheetnames.index('optimal')
    sheet = wb.active

    sql = analyzeSQL.optimal
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    # [Step 67 RESOURCE_LIMIT]
    sheet = wb.create_sheet('resource_limit')
    wb.active = wb.sheetnames.index('resource_limit')
    sheet = wb.active

    sql = analyzeSQL.resource_limit
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    max_cell_val = int(sheet.max_column) + 2

    # 0.process
    data_col = ["D", "E", "F"]
    cat_col = ["C"]
    title = "process"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 1.sessions
    data_col = ["G", "H", "I", "J"]
    cat_col = ["C"]
    title = "sessions"
    chart_num = 1
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 2.enqueue
    data_col = ["K", "L", "M"]
    cat_col = ["C"]
    title = "enqueue"
    chart_num = 2
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 3.parallel max servers
    data_col = ["Q", "R", "S"]
    cat_col = ["C"]
    title = "parallel max servers"
    chart_num = 3
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 68 MTTR_SQL]
    sheet = wb.create_sheet('mttr_sql')
    wb.active = wb.sheetnames.index('mttr_sql')
    sheet = wb.active

    sql = analyzeSQL.mttr_sql
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    max_cell_val = int(sheet.max_column) + 2

    # 0.Estimated MTTR(s)
    data_col = ["D"]
    cat_col = ["C"]
    title = "Estimated MTTR(s)"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 69 SGA]
    sheet = wb.create_sheet('sga')
    wb.active = wb.sheetnames.index('sga')
    sheet = wb.active

    sql = analyzeSQL.sga
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    max_cell_val = int(sheet.max_column) + 2

    # 0.SGA factor size(MB)
    data_col = ["D", "F", "E", "G", "H", "I"]
    cat_col = ["C"]
    title = "SGA factor size(MB)"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 1.Shared pool factor size(MB)
    data_col = ["J", "K", "L", "M"]
    cat_col = ["C"]
    title = "Shared pool factor size(MB)"
    chart_num = 1
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 2.Large Pool usage(MB)
    data_col = ["N", "O"]
    cat_col = ["C"]
    title = "Large Pool usage(MB)"
    chart_num = 2
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 3.Java pool usage(MB)
    data_col = ["P", "Q"]
    cat_col = ["C"]
    title = "Java pool usage(MB)"
    chart_num = 3
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 4.Streams pool usage(MB)
    data_col = ["R", "S"]
    cat_col = ["C"]
    title = "Streams pool usage(MB)"
    chart_num = 4
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 70 SGA_ROW]
    sheet = wb.create_sheet('sga_row')
    wb.active = wb.sheetnames.index('sga_row')
    sheet = wb.active

    sql = analyzeSQL.sga_raw
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    # [Step 71 BUFPOOL]
    sheet = wb.create_sheet('bufpool')
    wb.active = wb.sheetnames.index('bufpool')
    sheet = wb.active

    sql = analyzeSQL.bufpool
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    max_cell_val = int(sheet.max_column) + 2

    # 0.buffer gets
    data_col = ["G"]
    cat_col = ["C"]
    title = "buffer gets"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 1.free buffer waits
    data_col = ["J"]
    cat_col = ["C"]
    title = "free buffer waits"
    chart_num = 1
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 2.write complete waits
    data_col = ["K"]
    cat_col = ["C"]
    title = "write complete waits"
    chart_num = 2
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 3.buffer busy waits
    data_col = ["L"]
    cat_col = ["C"]
    title = "buffer busy waits"
    chart_num = 3
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 72 ROWCACHE]
    sheet = wb.create_sheet('rowcache')
    wb.active = wb.sheetnames.index('rowcache')
    sheet = wb.active

    sql = analyzeSQL.rowcache
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    # [Step 73 LIBRARYCACHE]
    sheet = wb.create_sheet('librarycache')
    wb.active = wb.sheetnames.index('librarycache')
    sheet = wb.active

    sql = analyzeSQL.librarycache
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)
        max_cell_val = int(sheet.max_column) + 2

    # 0.Library cache reloads
    data_col = ["D", "E", "F", "G"]
    cat_col = ["C"]
    title = "Library cache reloads"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 1.Library cache invalidations
    data_col = ["H", "I", "J", "K"]
    cat_col = ["C"]
    title = "Library cache invalidations"
    chart_num = 1
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 74 LIBCACHE_RAW]
    sheet = wb.create_sheet('libcache_raw')
    wb.active = wb.sheetnames.index('libcache_raw')
    sheet = wb.active

    sql = analyzeSQL.libcache_raw
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    # [Step 75 SESS_CACHE]
    sheet = wb.create_sheet('sess_cache')
    wb.active = wb.sheetnames.index('sess_cache')
    sheet = wb.active

    sql = analyzeSQL.sess_cache
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)
    max_cell_val = int(sheet.max_column) + 2

    # 0.session cached cursor hit(%)
    data_col = ["G"]
    cat_col = ["C"]
    title = "session cached cursor hit(%)"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 76 UNDO]
    sheet = wb.create_sheet('undo')
    wb.active = wb.sheetnames.index('undo')
    sheet = wb.active

    sql = analyzeSQL.undo
    cursor.execute(sql, snap_start=vsnap_start, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)
    max_cell_val = int(sheet.max_column) + 2

    # 0.undo blocks(k)
    data_col = ["E"]
    cat_col = ["D"]
    title = "undo blocks(k)"
    chart_num = 0
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 1.transaction count
    data_col = ["F"]
    cat_col = ["D"]
    title = "transaction count"
    chart_num = 1
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 2.Max Query Length(s)
    data_col = ["G"]
    cat_col = ["D"]
    title = "Max Query Length(s)"
    chart_num = 2
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 3.snapshot too old count
    data_col = ["K"]
    cat_col = ["D"]
    title = "snapshot too old count"
    chart_num = 3
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # 4.out of space count
    data_col = ["L"]
    cat_col = ["D"]
    title = "out of space count"
    chart_num = 4
    LineChart(data_col, cat_col, title, sheet, max_cell_val, chart_num)

    # [Step 77 PARAMETER]
    sheet = wb.create_sheet('parameter')
    wb.active = wb.sheetnames.index('parameter')
    sheet = wb.active

    sql = analyzeSQL.parameter
    cursor.execute(sql, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    # [Step 78 SGA_PARAM]
    sheet = wb.create_sheet('sga_param')
    wb.active = wb.sheetnames.index('sga_param')
    sheet = wb.active

    sql = analyzeSQL.sga_param
    cursor.execute(sql, snap_end=vsnap_end, dbid=vdbid, instno=vinstno)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    # [Step 79 GEN_RPT2]
    sheet = wb.create_sheet('gen_rpt2')
    wb.active = wb.sheetnames.index('gen_rpt2')
    sheet = wb.active

    sql = analyzeSQL.gen_rpt2
    cursor.execute(sql)

    columns = [desc[0] for desc in cursor.description]
    data = cursor.fetchall()

    df = pd.DataFrame(list(data), columns=columns)
    df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)


    try:
        wb.save(path + "\\" + file_name + ".xlsx")
        result = "분석 완료"
        return result
    except cx_Oracle.DatabaseError as err:
        return str(err)
    finally:
        cursor.close()
        conn.close()



